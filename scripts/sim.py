import os
from pathlib import Path
import hashlib
from  openpyxl import load_workbook
from openpyxl.utils import get_column_letter
from rosetta_sip_factory.sip_builder import build_sip_from_json
import time
import json
import shutil
import magic #python_magic_bin
import gzip
from datetime import datetime as dt
import logging


logger = logging.getLogger(__name__)
timestring = dt.now().strftime("_%Y_%m_%d")
path = r"Y:\ndha\pre-deposit_prod\LD_working\SIM"
#sprsh_path = r"D:\my_bool.xlsx"
names = ['Aaron', 'Andrea', 'Celeste', 'Diana', 'Gavin', 'Justin', 'Kim', 'Lynley', 'Maria', 'Melissa', 'Michelle', 'Rhonda', 'Sian', 'Silvia', 'Theresa', 'Tine']
#names = ["Melissa"]
ws_names  = ["continuous","one-time","warc"]
#ws_names  = ["warc_test"]
script_dir = os.getcwd()
project_dir = str(Path(script_dir).parents[0])
log_dir = os.path.join(r"Y:\ndha\pre-deposit_prod\LD_working\SIM\z_logs")
error_dir = os.path.join(log_dir,"errors")
completed_dir = os.path.join(log_dir,"completed")
accets_dir = os.path.join(project_dir,"assets")
sip_dir = os.path.join(project_dir,"sip")
dirs = [error_dir, completed_dir, accets_dir,sip_dir]
for dr in dirs:
	if not os.path.isdir(dr):
		os.makedirs(dr)
error_file = os.path.join(log_dir,"errors",f"errors{timestring}.txt")
completed_files = os.path.join(log_dir,"completed",f"completed{timestring}.txt")
rosetta_folder =  r"Y:\ndha\pre-deposit_prod\server_side_deposits\prod\ld_scheduled"
rosetta_periodic = os.path.join(rosetta_folder,"periodic")
rosetta_periodic_audio_video =os.path.join(rosetta_folder, "periodic_audio_video")
rosetta_oneoff = os.path.join(rosetta_folder,"oneoff")
rosetta_oneoff_audio=os.path.join(rosetta_folder,"oneoff_audio")
rosetta_oneoff_video=os.path.join(rosetta_folder,"oneoff_video")
rosetta_warc = os.path.join(rosetta_folder,"Warc")
entity_types = {"One Time":"OneOffIE","Audio (one time)":"AudioIE", "Continuous": "PeriodicIE", "Video (one time)" :"VideoIE","Warc - HTML Serial":"HTMLSerialIE","Warc - HTML Mono":"HTMLMonoIE"}
main_log = os.path.join(log_dir,"completed",f"main_log{timestring}.txt")


class SIPMaker():

	
	# Volume = dcterms:bibliographicCitation
	# Issue = dcterms:issued
	# Number = dcterms:accrualPeriodicity
	# Year = dc:date
	# Month = dcterms:available
	# Day = dc:coverage
	def __init__(self,descript, my_dict, files_path):

		"""This class is making sips for periodic, oneoff and warc workflow


		Variables:


		Methods:

		generate_sips -build sips from jsons for periodic, oneoff, warc
		make_json - build single dictionary with metadata

		"""
	
		self.my_dict = my_dict
		self.descript = descript
		self.count_done = 0
		self.count_failed = 0
		self.count_all = 0
		
		self.json_list = []
		#print(self.my_dict)
		try:
			for ln in self.my_dict:
				self.entity = ln["entity_type"]
				self.access = ln["access"]
				self.title = ln["title"]
				self.mmsid = ln["mmsid"]
				self.volume = ln["volume"]
				self.primary_url = ln["primary_url"]
				self.sprsh_path = ln["sprsh_path"]
				self.workflow = ln["workflow"]
				self.harvest_date = ln["harvest_date"]

				if self.volume =="None":
					self.volume = ""
				self.number = ln["number"]
				if self.number == "None":
					self.number = ""
				self.issue = ln["issue"]
				if self.issue == "None":
					self.issue = ""
				self.year = ln["year"]
				if self.year == "None":
					self.year = ""
				self.day = ln["day"]
				if self.day == "None":
					self.day = ""
				self.month = ln["month"]
				if self.month == "None":
					self.month = ""
				self.filepath = ln["filepath"]
				self.file_folder = os.path.dirname(self.filepath)
				self.filename= os.path.basename(self.filepath)
				self.files_path = files_path
				if ln["label"]:
					self.label = ln["label"]
				else:
					self.label = self.filename.split(".")[0]
				if self.label == "None":
					self.label = self.filename.split(".")[0]
				one_json = self.make_json()
				self.json_list =self.json_list+[one_json]
			self.ie_dc_dict = [{"dcterms:bibliographicCitation":self.volume,"dcterms:accrualPeriodicity":self.number,"dcterms:issued":self.issue,"dc:date":self.year,"dcterms:available":self.month,"dc:coverage": self.day,"dc:title":self.title}]
			self.input_dir = self.file_folder
			self.pres_master_json = json.dumps(self.json_list)
			self.general_ie_chars=[{'IEEntityType':self.entity}]
			self.web_harvesting = [{"primarySeedURL":self.primary_url,"harvestDate":self.harvest_date,"WCTIdentifier":'Webrecorder'}]
			self.object_identifier=[{'objectIdentifierType': 'ALMAMMS', 'objectIdentifierValue': self.mmsid}] 
			self.access_rights_policy=[{'policyId': self.access}]   
			self.sip_title = "SIM_"+self.descript
			self.output_folder = os.path.join(sip_dir, self.sip_title)
			self.kwargs = {'ie_dmd_dict':self.ie_dc_dict, 
					'generalIECharacteristics':self.general_ie_chars,
					'objectIdentifier':self.object_identifier,
					'accessRightsPolicy':self.access_rights_policy,
					'input_dir': self.input_dir,
					'digital_original':True,
					'sip_title':self.sip_title,
					'output_dir':self.output_folder,
					'encoding':'utf-8'}


		except Exception as e:
			print(str(e))
			self.count_all +=1
			self.count_failed +=1
			with open(error_file,"a") as f:
				f.write(str(e)+"\n")
				f.write(self.sprsh_path+"|"+self.workflow+"|"+self.descript+"|"+self.title+"\n")
	
	def build_sip_from_folder(self):

		logger.info("Making sips")

		if self.workflow == "warc":
			self.kwargs ['webHarvesting']=self.web_harvesting
		try:
			build_sip(**self.kwargs)
			print(self.descript, self.title, self.entity,"processing")

			with open(completed_files,"a") as f:
				f.write(self.sprsh_path+"|"+self.workflow+"|"+self.descript+"|"+self.title+"|"+self.entity+"\n")
			return 1
		except Exception as e:

			with open(error_file,"a") as f:
				f.write(str(e)+"\n")
				f.write(self.sprsh_path+"|"+self.workflow+"|"+self.descript+"|"+self.title+"\n")
			return 0

	def make_json(self):

		"""Makes json dictionary with mets metadata

		Returns:
			my_json(dict)-contains mets metadata
		"""
		
		my_json = {}
		my_json['physical_path'] = self.filepath
		my_json["fileOriginalName"]= self.filename
		my_json["fileOriginalPath"] = self.filename
		my_json["MD5"] = make_fixity(self.filepath)
		my_json["fileSizeBytes"] = str(os.path.getsize(self.filepath))
		my_json["fileCreationDate"] = time.strftime("%Y-%m-%dT%H:%M:%S",time.localtime(os.path.getctime(self.filepath)))
		my_json["fileModificationDate"] = time.strftime("%Y-%m-%dT%H:%M:%S",time.localtime(os.path.getmtime(self.filepath)))
		my_json["label"] = self.label

		return my_json

		
	def generate_sips(self):

		"""
		Build SIP folder from json list for warcs and non  warcs

		"""

		self.count_all+=1

	
		#print(access_rights_policy)
		try:
			self.kwargs["pres_master_json"]=self.pres_master_json
		except:
			print("File is not found in given location")
			quit()
		if self.workflow == "warc":
			self.kwargs ['webHarvesting']=self.web_harvesting
			
		try:
				build_sip_from_json(**self.kwargs)
	
				print(self.descript, self.title, self.entity,"processing")
				self.count_done +=1
				with open(completed_files,"a",encoding = "utf-8") as f:
					f.write(self.sprsh_path+"|"+self.workflow+"|"+self.descript+"|"+self.title+"|"+self.entity+"\n")

		except Exception as e:

				print(str(e))
				self.count_failed+=1
				with open(error_file,"a", encoding = "utf-8") as f:
					f.write(str(e)+"\n")
					f.write(self.sprsh_path+"|"+self.workflow+"|"+self.descript+"|"+self.title+"\n")
		

def sip_checker(sippath):

	"""Checks if met files are empty, or no_file
		Parameters:
		sippath(str) - path to sip
		Returns:
		flag(bool) - True if error found.  False if size of file is wrong or audio file or met file are empty.
	"""
	flag = False

	if os.path.getsize(os.path.join(sippath, "content", "mets.xml")) == 0:
		logger.info("Attention - empty met! {} ".format(sippath))
		flag = True
	if os.path.getsize(os.path.join(sippath, "content", "dc.xml")) == 0:
		logger.info("Attention - empty  dc met! {}".format(sippath))
		flag = True
	if len(os.listdir(os.path.join(sippath,  "content", "streams"))) == 0:
		logger.info("Attention - no file! {}".format(sippath))
		flag = True
	if len(os.listdir(os.path.join(sippath,  "content"))) == 0:
		logger.info("Attention - streem folder! {}".format(sippath))
		flag = True
	else:
		myfilepath = os.path.join(sippath, "content", "streams", os.listdir(os.path.join(sippath,  "content", "streams"))[0])
		if os.path.getsize(myfilepath) == 0:
				logger.info("Attention - 0 byte file! {}".format(myfilepath))
				flag = True				
	return flag

def make_fixity(f):

        BLOCKSIZE = 65536
        hasher = hashlib.md5()
        with open(f, 'rb') as afile:
                buf = afile.read(BLOCKSIZE)
                while len(buf) > 0:
                    hasher.update(buf)
                    buf = afile.read(BLOCKSIZE)
        return hasher.hexdigest()

def gzip_process (filepath):

	"""Unzips warc and other files
	"""

	parent_path = str(Path(filepath).parents[0])
	temp_path = os.path.join(parent_path, "temp.warc")
	shutil.move(filepath, temp_path)
	try:
		with gzip.open(temp_path, 'rb') as f_in:
			with open(filepath, 'wb') as f_out:
				shutil.copyfileobj(f_in, f_out)
	except:
		quit()

class SIM_spreadsheet():

	def __init__(self):

		"""This class is core of SIM process"""

		self.my_dict = None

	def read_spreadsheet(self,sprsh_path):

		"""Reading spreadsheet metadata
		Parameters:
			sprsh_path(str) - path to spreadsheet
		Returns:
			my_dict(dict) - contains spreadsheet metadata gathered by description

		"""


		wb = load_workbook(sprsh_path)
		self.my_dict = {}
		all_sprsh_rows = 0
		for sheet_name in ws_names:

			ws = wb[sheet_name]
			#print(ws.cell(2,2).value)
			for row in range(2, ws.max_row+1):

				
				if len(str(ws.cell(row,2).value))<10:
					break
				# fields = ["filepath", "mmsid", "title", "volume", "number", "year", "month", "day","access", "entity_type", "label", "primary_url", "harvest_date", "sheet_name"]
				# for fld in fields:
				# 	globals()[fld] = None
				all_sprsh_rows+=1
				filepath = None
				mmsid= None
				title= None
				volume= None
				issue = None
				number= None
				year= None
				month= None
				day= None
				access= None
				entity_type= None
				label= None
				primary_url= None
				harvest_date= None
				my_sheet_name= None
				folder_count = 0
				folder_count_all = 0

				if sheet_name == "continuous":
					filepath=str(ws.cell(row,1).value)
					mmsid=str(ws.cell(row,2).value)
					title = str(ws.cell(row,3).value)
					volume=str(ws.cell(row,4).value)
					number=str(ws.cell(row,5).value)
					issue=str(ws.cell(row,6).value)
					year=str(ws.cell(row,7).value)
					month=str(ws.cell(row,8).value).zfill(2)
					day=str(ws.cell(row,9).value).zfill(2)
					access=str(ws.cell(row,10).value)
					entity_type=entity_types[str(ws.cell(row,11).value)]
					label =str(ws.cell(row,12).value)
				if sheet_name == "one-time":
					filepath=str(ws.cell(row,1).value)
					mmsid=str(ws.cell(row,2).value)
					title =str(ws.cell(row,3).value)
					access=str(ws.cell(row,4).value)
					entity_type=str(entity_types[str(ws.cell(row,5).value)])
					label=str(ws.cell(row,6).value)
				if sheet_name == "warc":
					filepath=str(ws.cell(row,1).value)
					mmsid=str(ws.cell(row,2).value)
					title = str(ws.cell(row,3).value)
					primary_url=str(ws.cell(row,4).value)
					harvest_date=str(ws.cell(row,5).value)
					volume=str(ws.cell(row,6).value)
					number=str(ws.cell(row,7).value)
					issue=str(ws.cell(row,8).value)
					year=str(ws.cell(row,9).value)
					month=str(ws.cell(row,10).value).zfill(2)
					day=str(ws.cell(row,11).value).zfill(2)
					access=str(ws.cell(row,12).value)
					entity_type=entity_types[str(ws.cell(row,13).value)]
					label =str(ws.cell(row,14).value)

				my_sheet_name = str(sheet_name)
				fields = [filepath, title, mmsid, volume, number, issue, year, month, day, access, entity_type, label, primary_url, harvest_date, sheet_name]
				for i, el in enumerate(fields):
					if  el == 'None':
						fields[i] = ""
					elif not el:
						fields[i] = ""
					else:
						fields[i] = str(el)
				my_ie = "_".join(fields[2:9]).replace("__","_").replace("/","_").rstrip("_")


				if not  os.path.isdir(filepath):
					mime = magic.Magic(mime=True)
					if my_sheet_name == "warc" and "gzip" in mime.from_file(filepath):
						gzip_process(filepath)
					if not my_ie  in self.my_dict.keys():
						self.my_dict[my_ie]=[{"filepath":filepath, "title":title, "mmsid":mmsid, "volume":volume, "number":number, "issue":issue, "year":year, "month":month, "day":day, "access":access, "entity_type":entity_type, "label":label, "primary_url":primary_url, "harvest_date":harvest_date,"workflow":sheet_name,"sprsh_path":sprsh_path}]
					else:
						self.my_dict[my_ie]+=[{"filepath":filepath, "title":title, "mmsid":mmsid, "volume":volume, "number":number, "issue":issue, "year":year, "month":month, "day":day, "access":access, "entity_type":entity_type, "label":label, "primary_url":primary_url, "harvest_date":harvest_date,"workflow":sheet_name,"sprsh_path":sprsh_path}]


				else:
					small_dict = {}
					folder_count_all+=1
					my_files = os.listdir(filepath)
					for fl in my_files:
						filename = os.path.join(filepath, fl)
						mime = magic.Magic(mime=True)
						if my_sheet_name == "warc" and "gzip" in mime.from_file(filepath):
							gzip_process(filepath)
						small_dict[my_ie]=[{"filepath":filepath, "title":title, "mmsid":mmsid, "volume":volume, "number":number, "issue":issue, "year":year, "month":month, "day":day, "access":access, "entity_type":entity_type, "label":label, "primary_url":primary_url, "harvest_date":harvest_date,"workflow":sheet_name,"sprsh_path":sprsh_path}]
						small_sip =SIPMaker(ie, small_dict[ie],filepath)
						my_count = small_sip.build_sip_from_folder()
						folder_count+=my_count
	
		return self.my_dict, folder_count_all, folder_count, all_sprsh_rows


def sim_routine():

		"""
		Managing spreadsheet walking, gathering metadata, passing to SIPMaker,
	 	writing reports, moving files
	 	"""

		for name in os.listdir(path):
			named_folder = os.path.join(path,name)
			if os.path.isdir(named_folder) and name in names:
				print("#"*50)
				print("Name: ", name)
				sprsh_path = os.path.join(path,name,"ready_spreadsheets")
				processed_sprsh_path = os.path.join(path,name,"processed")
				files_path = os.path.join(path,name,"files" )
				processed_files_path = os.path.join(path,name, "processed", "files")
				my_folders = [sprsh_path, processed_sprsh_path, files_path, processed_files_path, sip_dir]
				for folder in my_folders:
					if not os.path.isdir(folder):
						os.makedirs(folder)
				sprsh_done =0
				sprsh_failed =0
				from_folder =0 
				for sprsh in os.listdir(sprsh_path):
					my_sim = SIM_spreadsheet()
					if not sprsh.startswith("~") and not sprsh.endswith(".xltm"):
						print("Spreadsheet:", sprsh)
						one_sprsh_path = os.path.join(sprsh_path, sprsh)
						dictionaries, folder_count_all, folder_count, all_sprsh_rows = my_sim.read_spreadsheet(one_sprsh_path)
						print("number of IEs submitting " ,len(dictionaries))
						print("number of rows ", all_sprsh_rows)
						for descript in dictionaries.keys():
							my_sip = SIPMaker(descript, dictionaries[descript], files_path)
							my_sip.generate_sips()
							sprsh_done+=my_sip.count_done
							sprsh_failed+=my_sip.count_failed
						print("Done:", sprsh_done + folder_count)
						print("Failed: ",sprsh_failed +(folder_count_all - folder_count))
						for sip in os.listdir(sip_dir):
							my_sip_folder = os.path.join(sip_dir, sip)
							dict_key = sip.lstrip("SIM_")
							flag = sip_checker(my_sip_folder)
							# print(flag)
							# print(len(dictionaries))
							if not flag and sprsh_failed == 0:
								my_entity_type=dictionaries[dict_key][0]["entity_type"]
								if my_entity_type == "PeriodicIE":
									fold = "periodic"
									shutil.move(my_sip_folder, rosetta_periodic)
								elif my_entity_type == "OneOffIE":
									fold = "oneoff"
									shutil.move(my_sip_folder, rosetta_oneoff)

								elif my_entity_type == "AudioIE":
									fold = "oneoff_audio"
									shutil.move(my_sip_folder, rosetta_oneoff_audio)
								elif my_entity_type == "VideoIE":
									fold = "oneoff_video"
									shutil.move(my_sip_folder, rosetta_oneoff_video)
								elif my_entity_type == "PeriodicAudio":
									fold = "periodic_audio_video"
									shutil.move(my_sip_folder, rosetta_periodic_audio_video)
								elif my_entity_type == "HTMLSerialIE":
									fold = "Warc"
									shutil.move(my_sip_folder, rosetta_warc)
								elif my_entity_type == "HTMLMonoIE":
									fold = "Warc"
									shutil.move(my_sip_folder, rosetta_warc)
								print(sip, "moved to", fold)
						try:
							shutil.move(one_sprsh_path, processed_sprsh_path)
							print("Spreadsheet "+one_sprsh_path+" moved to " + processed_sprsh_path)
						except Exception as e:
							print(str(e))
							with open(error_file , "a",encoding = "utf-8") as f:
								f.write("Could not move spreadsheet "+sprsh_path)
								f.write("\n")
						for ie in dictionaries.keys():
							for el in dictionaries[ie]:
								try:
									file_or_folder = el["filepath"].split("\\")[-2]
									if file_or_folder == "files":
										shutil.move(el["filepath"], processed_files_path)
									else:
										if not os.path.isdir(os.path.join(processed_files_path, file_or_folder)):
											os.makedirs(os.path.join(processed_files_path, file_or_folder))
										shutil.move(el["filepath"], os.path.join(processed_files_path, file_or_folder))
									print(file_or_folder, "moved to ", processed_files_path)
								except Exception as e:
									print(str(e))
						with open(main_log,"a",encoding = "utf-8") as f:
							f.write(name+'|'+ sprsh+"|" +str(len(dictionaries))+'|'+str(sprsh_done+folder_count) +'|'+ str(sprsh_failed+(folder_count_all-folder_count))+"\n")
 

def main():

	sim_routine()



if __name__ == '__main__':

    main()
